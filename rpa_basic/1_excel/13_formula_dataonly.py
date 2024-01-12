from openpyxl import load_workbook  # 파일 불러오기
# wb = load_workbook("sample_formula.xlsx")  # sample_formula.xlsx 파일에서 wb 을 불러옴
# ws = wb.active  # 현재 활성화된 Sheet

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)  # sample_formula.xlsx 파일에서 wb 을 불러옴
ws = wb.active  # 현재 활성화된 Sheet

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# 엑셀 한번 실행하고 닫을때 저장을 해주면 None 이 아닌 값이 나옴
for row in ws.values:
    for cell in row:
        print(cell)
        
print(" < 모두 성공 적으로 완성되었습니다 > ")
