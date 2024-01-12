from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2")  # B2 부터 D2 까지 합치겠음
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")
print(" < 모두 성공 적으로 완성되었습니다 > ")
