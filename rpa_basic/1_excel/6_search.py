from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active # 현재 활성화된 워크시트 불러오기

for row in ws.iter_rows(min_row=2): # 2번째 줄부터
    # 번호, 영어, 수학
    if int(row[1].value) > 80: # row 괄호안의 숫자는 키값의 위치
        print(row[0].value, "번 학생은 영어 천재")
        
for row in ws.iter_rows(max_row=1): # 최대 1번째 줄까지(첫번째 줄만)
    for cell in row:                # for 반복문으로 row 행전체 검색     
        if cell.value == "영어": # 영어를 컴퓨터로 변경하기
            cell.value = "컴퓨터"
                
wb.save("sample_modified.xlsx")