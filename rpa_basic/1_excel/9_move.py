from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

# 번호 영어 수학
# 번호 (국어) 영어 수학
# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어" # B1 셀에 "국어" 입력

# 번호 영어 수학  
# C의 11줄이 영어의 5번째 줄부터 수학이 왼쪽으로(-1) 이동하여 덮어씀 
ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")
print(" < 모두 성공 적으로 완성되었습니다 > ")