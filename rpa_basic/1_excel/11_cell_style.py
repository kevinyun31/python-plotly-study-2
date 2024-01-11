from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("sample.xlsx")  # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨갛게,이탤릭, 굵게
b1.font = Font(color="cc33ff", name="Arial", strike=True) # 폰트를 Arial 로 설정, 취소선
c1.font = Font(color="0000ff", size=20, underline="single") # 글자 크기를 20으로, 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin") )
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:  # 전체 row
    for cell in row:  # 전체 column
        # Alignment = 각 cell 에 대해서 정렬
        # center, left, right, top, bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if cell.column == 1:  # A 번호열은 제외
            continue
        
        # cell 이 정수형 데이터이고 90 점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90 :
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid")  # 배경을 초록색으로 설정
            cell.font = Font(color="ff0000")  # 폰트 색상 변경       
        
# 틀 고정
ws.freeze_panes = "B2"  # B2 기준으로 틀 고정

wb.save("sample_style.xlsx")
print(" < 모두 성공 적으로 완성되었습니다 > ")
