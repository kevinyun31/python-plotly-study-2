import datetime
from openpyxl import Workbook  # 파일 새로 만들기
wb = Workbook()  # 새로운 워크북 객체
ws = wb.active # 활성화된 Sheet

ws["A1"] = datetime.datetime.today()  # 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)"  # 1 + 2+ 3 = 6 (합계)
ws["A3"] = "=AVERAGE(1, 2, 3)"  # 6 (평균)

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"  # 30

wb.save("sample_formula.xlsx")
print(" < 모두 성공 적으로 완성되었습니다 > ")
