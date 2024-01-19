import openpyxl

fpath = r'C:\workspace\python_lab\STARTCODING\02_파이썬엑셀다루기\참가자_data.xlsx'

# 1. 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2. 엑셀 시트선택
ws = wb['오징어게임']

# 3. 데이터 수정하기
ws['A3'] = 456
ws['A4'] = 138
ws['B3'] = '성기훈'
ws['B4'] = '정우성'

# 4. 데이터 삭제하기
ws.delete_rows(4)

# 5. 엑셀 저장하기
wb.save(fpath)
print(" <모든 작업이 성공적으로 완료되었습니다> ")