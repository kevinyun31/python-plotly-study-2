import openpyxl
from random import randint
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter

# 1. 엑셀 만들기
wb = openpyxl.Workbook()

# 2. 엑셀 워크시트 만들기
ws = wb.active
ws.title = "주식데이터"  # 주어진 이름으로 Sheet 이름 변경 

# 3. 데이터 추가하기
ws.append(['종목', '현재가', '평균매입가', '잔고수량', '평가금액', '평가손익', '수익률'])

# columns = 반복문으로 한 행씩 넣어야 함
data_columns = ['삼성전자', 'SK하이닉스', '카카오', '제주반도체' ,'POSCO홀딩스',
                '셀트리온', 'NAVER', '포스코DX', '에코프로', '포스코인터내셔널']
for company in data_columns:
    ws.append([company, ''])

# 4. 랜덤 숫자 채우기
for x in range(2, 12):  # 2~11까지 10개 row
    for y in range(3, 4):  # B column
        ws.cell(row=x, column=y, value=randint(80000, 200000))  # 80000~200000 사이의 숫자 

for x in range(2, 12):  # 2~11까지 10개 row
    for y in range(4, 5):  # C column
        ws.cell(row=x, column=y, value=randint(10, 30))  # 10~30 사이의 숫자 

# 5. 평가금액(=B * D) 및 평가손익(E - C*D) 열에 계산식 추가
for x in range(2, 12):  # 2~11까지 10개 row
    ws.cell(row=x, column=5, value=f'=B{x}*D{x}')
    ws.cell(row=x, column=6, value=f'=E{x}-C{x}*D{x}')
    # 수익률 추가
    ws.cell(row=x, column=7, value=f'=(B{x}-C{x})/C{x}')

# 6. 셀 형식 지정
# 천단위 "," 삽입
comma_style = NamedStyle(name='comma', number_format='#,##0')
for col_letter in ['B', 'C', 'D', 'E', 'F']:
    for row_num in range(2, 12):
        ws[f'{col_letter}{row_num}'].style = comma_style

# "%" 값으로 변경
percent_style = NamedStyle(name='percent', number_format='.00%')
for row_num in range(2, 12):
    ws[f'G{row_num}'].style = percent_style 


# 7. 셀 간격 조절
# 셀안의 값의 크기에 맞춰서 셀 간격 조절
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    max_length = 0
    for cell in row:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    for cell in row:
        ws.column_dimensions[get_column_letter(cell.column)].width = adjusted_width
       
# 8. 셀 정렬
# 전체를 중앙 정렬
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# B2에서 G11까지의 int 값 범위를 우측 정렬
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=7):
    for cell in row:
        cell.alignment = Alignment(horizontal='right')
        
# 9. 엑셀 저장하기   
wb.save(r'C:\workspace\python_lab\STARTCODING\01_네이버_주식현재가_크롤링\data.xlsx')
print("<모든 작업이 성공적으로 완료되었습니다>")
