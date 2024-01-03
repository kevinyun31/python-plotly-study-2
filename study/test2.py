#팬더 및 플롯 설치 - 아래 코드는 콘솔창에 입력한다
# pip install pandas
# pip install plotly

# --------------------------------------------------------------------------

#라이브러리를 가져오고 사용하기 쉽도록 별칭을 지정합니다.
import pandas as pd

import plotly.express as px

import plotly.io as pio

# --------------------------------------------------------------------------

#데이터세트를 읽고 변수 df에 저장
df = pd.read_excel('data2.xlsx')

# 총 매출을 계산하고 이에 대한 새 열을 추가합니다.
df['Total Sales'] = df['Units Sold'] * df['Price per unit']

# --------------------------------------------------------------------------

# 제품 대 총 판매량의 막대 차트를 만듭니다.
fig = px.bar(df, x='Product', y='Total Sales', title='Product Sales')

# 차트 영역의 테두리와 배경색을 설정합니다.
fig.update_layout(
    plot_bgcolor='white',
    paper_bgcolor='lightgray',
    width=500,
    height=300,
    shapes=[dict(type='rect', xref='paper',
            yref='paper',
            x0=0,
            y0=0,
            x1=1,
            y1=1,
            line=dict(
                color='black',
                width=2,
            ),
        )
    ]
)

#그래프를 표시하다
fig.show()

pio.write_json(fig, 'figure.json', pretty=True)

# 아래 코드 줄을 사용하여 막대 그래프를 이미지에 저장
print (f"이미지 생성")
pio.write_image(fig, 'bar2_graph.png')

print (f"이미지 저장")

# --------------------------------------------------------------------------

# 엑셀 파일 생성 및 데이터 프레임 저장
# ExcelWriter는 내부에 writer.save()를 포함하고 있다
excel_path = 'report2.xlsx'
with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
    df.to_excel(writer, index=False, sheet_name='Sales Data')

    # 이미지를 엑셀에 추가
    worksheet = writer.sheets['Sales Data']
    worksheet.insert_image('H1', 'bar2_graph.png')

print(f"Excel 파일이 생성 되었습니다 : {excel_path}")

# --------------------------------------------------------------------------

# 필요한 패키지와 Excel 보고서 파일을 가져오겠습니다.
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

# 기존 통합 문서를 로드하고 활성 워크시트를 선택합니다.
wb = load_workbook('report2.xlsx') 
ws = wb.active

# --------------------------------------------------------------------------

# 1. 정렬(Alignment):

# 모든 열의 너비를 20으로 설정
for col in ws.columns: 
    ws.column_dimensions[col[0].column_letter].width = 20

# 셀의 텍스트를 왼쪽 정렬로 설정합니다.
for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column): 
     for cell in row: 
         cell.alignment = Alignment(horizontal='left')

# 수정된 통합 문서를 Excel 파일에 저장
# 아래 코드는 중복되면 안됨 마지막에 한번만 실행되어야 함
# wb.save('report2.xlsx')
# print(f"Excel 파일이 수정 되었습니다 : {excel_path}")

# --------------------------------------------------------------------------

# 2. 테두리 추가(Adding borders):

# 새 통합 문서를 만들고 활성 워크시트를 선택합니다.
# wb = load_workbook('report2.xlsx') 
# ws = wb.active


# 특정 셀 범위 주위에 테두리를 설정합니다.
range_border = Border(left=Side(style='dashDot'),     
                       right=Side(style='dashed'), 
                       top=Side(style='double'),  
                       bottom=Side(style='hair'))

for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column): 
       for cell in row: 
         cell.border = range_border

# 수정된 통합 문서를 Excel 파일에 저장
# 아래 코드는 중복되면 안됨 마지막에 한번만 실행되어야 함
# wb.save('report2.xlsx')
# print(f"Excel 파일이 수정 되었습니다 : {excel_path}")

# --------------------------------------------------------------------------

# 3. 열강조 표시(Highlighting the columns)
# 열을 강조 표시하기 위한 채우기 패턴을 정의합니다.
fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')

# 시트의 첫 번째 열을 강조 표시합니다.
for col in ws.iter_cols(min_col=1, max_col=1):
    for cell in col:
        cell.fill = fill

# 수정된 통합 문서를 Excel 파일에 저장
# 아래 코드는 중복되면 안됨 마지막에 한번만 실행되어야 함
# wb.save('report2.xlsx')
# print(f"Excel 파일이 수정 되었습니다 : {excel_path}")

# --------------------------------------------------------------------------

# 4. 글꼴 스타일 추가(Adding font styles:)
# 첫 번째 행의 글꼴 스타일을 설정합니다.
from openpyxl.styles import Font, PatternFill
font = Font(name='Arial', size=12, bold=True, italic=True, color='FF0000')

# 첫 번째 행의 배경색을 설정합니다.
fill = PatternFill(start_color='23C4ED', end_color='23C4ED', fill_type='solid') 
for row in ws.iter_rows(min_row=1, max_row=1): 
    if row[0].row == 1: 
       for cell in row: 
           cell.fill = fill 
           cell.font = font

# 수정된 통합 문서를 Excel 파일에 저장
# 아래 코드는 중복되면 안됨 마지막에 한번만 실행되어야 함
# wb.save('report2.xlsx')
print(f"Excel 파일이 수정 되었습니다 : {excel_path}")

# --------------------------------------------------------------------------