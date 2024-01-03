# https://plotly.com/blog/automate-excel-reports-with-python/

# 이제 남은 것은 위 스크립트가 매일, 매주, 매월 또는 사용 사례에 관계없이 실행되도록 예약하는 것입니다. 
# 필요와 시나리오에 따라 Crontab, 작업 스케줄러, Automator 또는 Anacron을 사용할 수 있습니다.

# 모든 것을 Python 함수로 래핑하기

# 코드가 다루기 쉬울 뿐만 아니라 재사용이 가능하고 읽기 쉽도록 모든 코드를 각각의 Python 함수로 래핑해 보겠습니다.
     
import pandas as pd
import plotly.express as px
import plotly.io as pio
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


#스프레드시트 자동화 기능
def automate_excel_report(inputFilename, outputFilename):
    df = pd.read_excel(inputFilename)

    # Total Sales에 대한 새 열을 추가합니다.
    df['Total Sales'] = df['Units Sold'] * df['Price per unit']

    # Plotly를 사용하여 판매된 단위의 막대 그래프를 만듭니다.
    fig = px.bar(df, x='Product', y='Total Sales', title='Product Sales')

    # 차트 영역의 테두리와 배경색을 설정합니다.
    fig.update_layout(
        plot_bgcolor='white',
        paper_bgcolor='lightgray',

        shapes=[dict(
                type='rect',
                xref='paper',
                yref='paper',
                x0=0,
                y0=0,
                x1=1,
                y1=1,
                line=dict(color='black', width=2))])
    
    # 막대그래프를 이미지 파일로 저장
    pio.write_image(fig, 'bar2_graph.png')
    writer = pd.ExcelWriter(outputFilename, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sales Data')
    worksheet = writer.sheets['Sales Data']
    worksheet.set_column('F:F', 12)
    worksheet.insert_image('H1', 'bar2_graph.png')

    # 파일을 저장
    writer.close()

    # --------------------------------------------------------------------------

#보고서 자동화 기능
def automate_excel_formatting(inputFilename, outputFilename):
    # 엑셀 파일을 로드한다
    wb = load_workbook(inputFilename)
    ws = wb.active

    # 열을 강조 표시하는 채우기 패턴을 정의합니다.
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 시트의 열 'Product'을 강조표시합니다. 
    for cell in ws['A:A']:
        cell.fill = fill

        # =========================테두리======================
    # 특정 셀 범위 주위에 테두리를 설정합니다.
    range_border = Border(left=Side(style='medium'), 
                        right=Side(style='medium'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))

    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = range_border

    # =====================열 정렬====================
    # 모든 열의 너비를 20으로 설정
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    # 워크시트에 텍스트를 추가하고 왼쪽 정렬로 설정합니다.
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column):
        for cell in row:
            # cell.value = f'Text {cell.column}{cell.row}'
            cell.alignment = Alignment(horizontal='left')

    #=======================글꼴 스타일===============
    # 다른 행에 대한 글꼴 스타일을 설정합니다.
    font_name = Font(name='Times New Roman', bold=True)

    # 첫 번째 행의 배경색을 설정합니다.
    fill = PatternFill(start_color='23C4ED', end_color='23C4ED', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=1):
        if row[0].row == 1:
            for cell in row:
                cell.fill = fill
                cell.font = font_name

    # 통합 문서를 저장
    wb.save(outputFilename)

# 피벗 테이블 자동화 기능
def automate_excel_pivot_table(inputFilename, outputFilename):
    # Excel 파일을 pandas DataFrame에 로드합니다.
    # df = pd.read_excel("supermarket_sales5.xlsx")
    df = pd.read_excel(inputFilename)

    # 피벗 테이블을 생성합니다
    pivot_df = pd.pivot_table(df, values='Total',
                              index='Gender', columns='Payment', aggfunc='sum')

    # 피벗 테이블을 Excel 파일로 내보내기
    pivot_df.to_excel(outputFilename, sheet_name='Sheet1', index=True)

    # Plotly 그림을 만듭니다.
    fig = px.imshow(pivot_df)

    # 그림을 보여주세요
    fig.write_image('fig2.png')

    # --------------------------------------------------------------------------

# 이제 Python 함수를 호출하고 파일 이름을 전달하기만 하면 됩니다. 예:

#'data2.xlsx'라는 파일을 전달하는 중입니다. 입력으로 'report2.xlsx'라는 새 파일이 생성됩니다. 출력으로
automate_excel_report('data2.xlsx', 'report2.xlsx')
print(f"### data2.xlsx파일을 가져와서 report2.xlsx을 작성하였습니다.")

#'report2.xlsx' 전달 할 함수를 사용하여 생성된 파일은 'formattedReport2.xlsx'라는 새로운 형식의 Excel 파일을 생성합니다.
automate_excel_formatting('report2.xlsx', 'formattedReport2.xlsx')
print(f"### report2.xlsx파일을 가져와서 formattedReport2.xlsx을 작성하였습니다.")

#'pivot_dataset5.xlsx'라는 파일 전달 pivotTable2.xlsx라는 새 파일이 생성됩니다.
automate_excel_pivot_table('supermarket_sales5.xlsx', 'pivotTable2.xlsx')
print(f"### supermarket_sales5.xlsx파일을 가져와서 pivotTable2.xlsx을 작성하였습니다.")

# --------------------------------------------------------------------------

# 비올라! 모든 것이 당신을 위해 이루어졌습니다.
print(f"### 축하합니다! 모든 작업이 완료 되었습니다.")

# 이제 남은 유일한 작업은 특정 요구 사항에 따라 위에서 언급한 스크립트에 대한 일정을 매일, 매주, 매월 단위로 설정하는 것입니다. 
# 사용 사례에 따라 Crontab, 작업 스케줄러, Automator 또는 Anacron과 같은 도구를 사용할 수 있습니다.