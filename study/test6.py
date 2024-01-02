# https://plotly.com/blog/automate-excel-reports-with-python/

# Now, all that is left is to schedule the above script to run daily, weekly, monthly, or
# whatever your use case is. Depending on your needs and scenario, you can use the
# Crontab, Task Scheduler, Automator, or Anacron.

# Wrapping everything into Python functions

# Let's wrap all the code into respective Python functions so that the code not only
# becomes easy to handle but also reusable, and more readable.

import pandas as pd
import plotly.express as px
import plotly.io as pio
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

#Function for automating the spreadsheet
def automate_excel_report(inputFilename, outputFilename):
    df = pd.read_excel(inputFilename)

    # Add a new column for Total Sales
    df['Total Sales'] = df['Units Sold'] * df['Price per unit']

    # Create a bar graph of units sold using Plotly
    fig = px.bar(df, x='Product', y='Total Sales', title='Product Sales')

    # set the border and background color of the chart area
    fig.update_layout(
        plot_bgcolor='white',
        paper_bgcolor='lightgray',

        shapes=[dict(
                type='rect',
                xref='paper',
                yref='paper',
                x0=0,
                y0=0,
                x1=1,
                y1=1,
                line=dict(color='black', width=2))])

    # Save the bar graph as an image file
    pio.write_image(fig, 'bar1_graph.png')
    writer = pd.ExcelWriter(outputFilename, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sales Data')
    worksheet = writer.sheets['Sales Data']
    worksheet.set_column('F:F', 12)
    worksheet.insert_image('H1', 'bar1_graph.png')

    # save the file
    writer.close()

# --------------------------------------------------------------------------

#Function for automating the reports
def automate_excel_formatting(inputFilename, outputFilename):
    # # load the Excel file
    wb = load_workbook(inputFilename)
    ws = wb.active

    # # define the fill pattern to highlight the column
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # highlight the 'Product' column of the sheet
    for cell in ws['A:A']:
        cell.fill = fill

    # =========================Borders======================
    # set a border around a specific range of cells
    range_border = Border(left=Side(style='medium'), 
                        right=Side(style='medium'), 
                        top=Side(style='medium'), 
                        bottom=Side(style='medium'))

    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = range_border

    # =====================column alignmnet====================
    # set the width of all columns to 20
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    # add some text to the worksheet and set it left-aligned
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=ws.max_column):
        for cell in row:
            # cell.value = f'Text {cell.column}{cell.row}'
            cell.alignment = Alignment(horizontal='left')

    #=======================font styles===============
    # set font styles for different rows
    font_name = Font(name='Times New Roman', bold=True)

    # set background color for the first row
    fill = PatternFill(start_color='23C4ED', end_color='23C4ED', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=1):
        if row[0].row == 1:
            for cell in row:
                cell.fill = fill
                cell.font = font_name

    # save the workbook 
    wb.save(outputFilename)

# Function for automating the pivot table
def automate_excel_pivot_table(inputFilename, outputFilename):
    # Load the Excel file into a pandas DataFrame
    # df = pd.read_excel("supermarket_sales4.xlsx")
    df = pd.read_excel(inputFilename)

    # Create a pivot table
    pivot_df = pd.pivot_table(df, values='Total',
                              index='Gender', columns='Payment', aggfunc='sum')

    # Export the pivot table to an Excel file
    pivot_df.to_excel(outputFilename, sheet_name='Sheet1', index=True)

    # Create a Plotly figure
    fig = px.imshow(pivot_df)

    # Show the figure
    fig.write_image('fig1.png')

# --------------------------------------------------------------------------

# Now you just need to call the Python functions and pass the filename to it. Example:

#passing my file named 'data.xlsx' as an input will create a new file named 'report.xlsx' as an output
automate_excel_report('data1.xlsx', 'report1.xlsx')
print(f"### data1.xlsx파일을 가져와서 report1.xlsx을 작성하였습니다.")

#passing the 'report.xlsx' file generated using the above function will create a new formatted excel file named 'formattedReport.xlsx'
automate_excel_formatting('report1.xlsx', 'formattedReport1.xlsx')
print(f"### report1.xlsx파일을 가져와서 formattedReport1.xlsx을 작성하였습니다.")

#passing the file named 'pivot_dataset.xlsx' will create a new file named pivotTable.xlsx
automate_excel_pivot_table('supermarket_sales4.xlsx', 'pivotTable1.xlsx')
print(f"### supermarket_sales4.xlsx파일을 가져와서 pivotTable1.xlsx을 작성하였습니다.")

# --------------------------------------------------------------------------

# Viola! Everything is done for you.
print(f"### 축하합니다! 모든 작업이 완료 되었습니다.")

# Now, the only remaining task is to set a schedule for the script mentioned above, 
# whether it's on a daily, weekly, or monthly basis, depending on your specific needs.
# You can use tools such as Crontab, Task Scheduler, Automator, or Anacron as per your use case.



